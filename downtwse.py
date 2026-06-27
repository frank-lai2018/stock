import pandas as pd
import requests
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 關閉 SSL 憑證驗證的警告訊息
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def fetch_taiwan_stocks():
    # 加入 User-Agent 偽裝成一般 Google Chrome 瀏覽器，避免被政府防火牆阻擋
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }

    print("正在抓取上市股票資料...")
    url_twse = "https://openapi.twse.com.tw/v1/opendata/t187ap03_L"
    res_twse = requests.get(url_twse, verify=False, headers=headers) 
    
    # 防呆檢查：確認證交所回傳成功 (HTTP 200) 才繼續
    if res_twse.status_code != 200:
        print(f"❌ 上市資料抓取失敗！HTTP狀態碼: {res_twse.status_code}")
        return None
        
    df_twse = pd.DataFrame(res_twse.json())
    df_twse['市場別'] = '上市'

    print("正在抓取上櫃股票資料...")
    url_tpex = "https://www.tpex.org.tw/openapi/v1/mktdata/t187ap03_O"
    res_tpex = requests.get(url_tpex, verify=False, headers=headers) 
    
    # 防呆檢查：確認櫃買中心回傳成功 (HTTP 200) 才繼續
    if res_tpex.status_code != 200:
        print(f"❌ 上櫃資料抓取失敗！HTTP狀態碼: {res_tpex.status_code}")
        return None
        
    df_tpex = pd.DataFrame(res_tpex.json())
    df_tpex['市場別'] = '上櫃'

    # 合併上市與上櫃資料
    print("資料合併與清洗中...")
    df = pd.concat([df_twse, df_tpex], ignore_index=True)
    
    # 篩選我們需要的欄位並重新命名
    df = df[['公司代號', '公司簡稱', '市場別', '產業別']]
    df.columns = ['股票代碼', '中文名稱', '市場別', '產業別']

    # 將代碼轉為數字型態進行排序 (若為 KY 股等無法轉數字的會排在最後面)
    df['sort_key'] = pd.to_numeric(df['股票代碼'], errors='coerce')
    df = df.sort_values(by=['sort_key', '股票代碼']).drop(columns=['sort_key']).reset_index(drop=True)

    return df

def export_to_excel(df, output_path="Taiwan_Stock_List_Complete.xlsx"):
    print("正在匯出並美化 Excel 檔案...")
    wb = Workbook()
    ws = wb.active
    ws.title = "台股總表"

    # 寫入標題列
    headers = df.columns.tolist()
    ws.append(headers)

    # 設定色彩與邊框樣式
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    border_style = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    # 套用標題列美化
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style

    # 寫入資料並設定斑馬紋與置中對齊
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        ws.append(row)
        for c_idx in range(1, len(row) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style
            if r_idx % 2 == 0:
                cell.fill = zebra_fill

    # 調整欄寬
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25

    # 凍結第一列標題，並開啟自動篩選器
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 儲存檔案
    wb.save(output_path)
    print(f"✅ 完成！檔案已儲存為：{output_path}")

if __name__ == "__main__":
    stock_df = fetch_taiwan_stocks()
    if stock_df is not None:
        export_to_excel(stock_df)